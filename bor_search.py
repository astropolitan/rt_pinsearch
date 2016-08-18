# -*- coding: utf-8 -*-
"""
Created on Sat Jun 18 10:34:44 2016

@author: Madeleine

BOR_SEARCH.py

 Searches Board of Review records for a file of PINs and enters information
 about their appeal history.

 ***For a new project, highlight this word and alter indicated terms as needed:
             *CHANGE*

Modification history:
 6/18/16 - began coding
 6/22/16 - added inputs in for-loop for manually entering attorneys and
             results, printing to Excel workbook, added *CHANGE* statements
 6/23/16 - added option for No Data and option for continuing old list (check),
             added column to indicate whether the PIN packet is worth printing,
             started switching Excel writing to xlwt
"""

import xlwt

from openpyxl import Workbook

from rt_selenium import bor
from create_pin import create_pin

# ===========================================================================
#  Reading the PINs from a file and generating a list of PINs.
# ===========================================================================
filename = 'schaumburg16n.txt'                  # *CHANGE* for every list

pin_list = []
with open(filename) as file:
    for line in file:
        pin_list.append(line.strip().split(','))
pin_list = pin_list[1:]

pins = []
for p in range(len(pin_list)):
    pins += [create_pin(pin_list[p][0])]

choice = input('New PIN list? y/n ')


# ===========================================================================
#  Searching the PINs at CookCountyBoardOfReview.com
# ===========================================================================
results = []
for p in range(len(pins)):
    print(pins[p][1])
    segs = pins[p][0]
    browser = bor(segs, 1)
    pinstr = str(pins[p][1])

    res15_0 = input('2015 Attorney: ')
    if res15_0 == 'no data':
        result = [pinstr, 'No data', 'No data', 'No data', '*']
    else:
        if res15_0 == 'none':
            res15 = 'None'
        elif res15_0 == 'q':                      # allows quitting mid-loop
            print(pins[p][1])
            break
        else:
            res15_1 = input('Change? y/n ')
            if res15_1 == 'y':
                res15_1 = 'C - '
            else:
                res15_1 = 'N/C - '
            res15 = res15_1 + res15_0

        res14_0 = input('2014 Attorney: ')
        if res14_0 == 'none':
            res14 = 'None'
        else:
            res14_1 = input('Change? y/n ')
            if res14_1 == 'y':
                res14_1 = 'C - '
            else:
                res14_1 = 'N/C - '
            res14 = res14_1 + res14_0

        res13_0 = input('2013 Attorney: ')
        if res13_0 == 'none':
            res13 = 'None'
        else:
            res13_1 = input('Change? y/n ')
            if res13_1 == 'y':
                res13_1 = 'C - '
            else:
                res13_1 = 'N/C - '
            res13 = res13_1 + res13_0

        result = [pinstr, res13, res14, res15]
        if result[1:] == ['None', 'None', 'None']:
            result += ['*']

    results += [result]
    browser.close()

print('Data recording complete. Now converting to Excel file...')


# ===========================================================================
#  Creating an Excel workbook with the data:
# ===========================================================================
book = xlwt.Workbook()
ws = book.add_sheet(filename)

if choice == 'y':
    ws['A1'] = 'PIN'
    ws['B1'] = '2013 BoR'
    ws['C1'] = '2014 BoR'
    ws['D1'] = '2015 BoR'
    ws['E1'] = 'Print packet'

for i in range(len(results)):
    if results[i][-1] == '*':
        style = xlwt.easyxf('pattern: pattern solid, fore_color yellow')
    else:
        style = xlwt.easyxf()
    data = [x for x in results[i]]
    ws.write(0, 0, data, style)

book.save('schaumburg_results.xls')               # *CHANGE* for every list

print('Excel file complete.')
